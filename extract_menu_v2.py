import openpyxl
import json
from collections import defaultdict

files = {
    'ISTIKNANAH': r'c:\Users\dell\Documents\menu4\ISTIKNANAH FINAL MENU (2).xlsx',
    'SOHO': r'c:\Users\dell\Documents\menu4\soho menu---.xlsx'
}

def extract_menu(file_path):
    """Extract menu data from Excel file"""
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    menu_data = defaultdict(list)
    current_category = None
    skip_next = False
    
    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=1000), 1):
        values = [cell.value for cell in row]
        
        # Skip if this row should be skipped (e.g., description row)
        if skip_next:
            skip_next = False
            continue
        
        # Skip completely empty rows
        if not any(v is not None for v in values[:4]):
            continue
        
        col_a = values[0]
        col_b = values[1] if len(values) > 1 else None
        col_c = values[2] if len(values) > 2 else None
        
        # Skip header rows
        if col_a and isinstance(col_a, str):
            if col_a.strip() in ['الصنف', 'سعر البيع', 'السعر']:
                continue
        
        # Determine if this is a category or item
        is_category = col_b is None and isinstance(col_a, str)
        
        if is_category and col_a:
            # This is a category
            current_category = col_a.strip()
            if current_category not in menu_data:
                menu_data[current_category] = []
        elif current_category and col_a:
            # This is an item with name and price
            item_name = col_a.strip() if isinstance(col_a, str) else str(col_a)
            item = {"name": item_name}
            
            # Extract price
            if col_b and isinstance(col_b, (int, float)):
                item["price"] = float(col_b)
            elif col_b and isinstance(col_b, str):
                try:
                    item["price"] = float(col_b.replace(',', '').strip())
                except (ValueError, AttributeError):
                    pass
            
            # Check if there's a description in the next row
            # (Description rows typically have only text in col A, no price in col B)
            if idx < sheet.max_row:
                next_row = list(sheet.iter_rows(min_row=idx+1, max_row=idx+1, values_only=False))[0]
                next_values = [cell.value for cell in next_row]
                next_col_a = next_values[0]
                next_col_b = next_values[1] if len(next_values) > 1 else None
                
                # If next row has text in A but no price in B (or looks like a description)
                if next_col_a and isinstance(next_col_a, str) and (next_col_b is None or 
                    (isinstance(next_col_b, str) and not any(c.isdigit() for c in next_col_b[:3]))):
                    # Check if this looks like a continuation/description (usually longer text)
                    if len(str(next_col_a)) > 40 or next_col_a.count(' ') > 5:
                        item["desc"] = next_col_a.strip()
                        skip_next = True
            
            menu_data[current_category].append(item)
    
    return dict(menu_data)

# Extract both menus
all_menus = {}
for label, file_path in files.items():
    print(f"Extracting {label}...")
    all_menus[label] = extract_menu(file_path)

# Output as JSON
output = json.dumps(all_menus, ensure_ascii=False, indent=2)
print("\n" + "="*60)
print(output)

# Also save to file
with open(r'c:\Users\dell\Documents\menu4\menus.json', 'w', encoding='utf-8') as f:
    f.write(output)
print("\n" + "="*60)
print("Saved to menus.json")
