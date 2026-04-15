import openpyxl
import json

# Load SOHO menu from Excel
wb = openpyxl.load_workbook('soho menu---.xlsx')
ws = wb.active

soho_items = {}
current_category = None
pending_description = None

for i, row in enumerate(ws.iter_rows(values_only=True)):
    item_name = row[0]
    price = row[1] if len(row) > 1 else None
    
    if not item_name:
        continue
    
    # Check if this is a category (has text but no price, or ends with ":"
    if item_name and not price and (not isinstance(price, (int, float)) or price is None):
        # Could be category or description
        if item_name.endswith(':') or item_name in [
            'المشروبات الباردة', 'الشوربة', 'السلطات', 'المقبلات', 
            'الساندوتشات', 'ركن الباستا', 'الاطباق الرئيسية', 'السايد',
            'وجبات الاطفال', 'اكسترا', 'الصوصات', 'الحلويات',
            'المشروبات الساخنه', 'عصائر فريش', 'مشروبات صحيه',
            'ايس درينك', 'ميلك شيك', 'مشروبات مميزة', 'صودا كوكتيل', 'سموزى'
        ]:
            current_category = item_name.rstrip(':')
            if current_category not in soho_items:
                soho_items[current_category] = []
            pending_description = None
        else:
            # This is a description for the next item
            pending_description = item_name
    elif price and isinstance(price, (int, float)):
        # This is an item with price
        if current_category:
            soho_items[current_category].append({
                'name': item_name,
                'price': int(price),
                'desc': pending_description or ''
            })
            pending_description = None

# Load ISTIKNANAH menu from Excel
wb2 = openpyxl.load_workbook('ISTIKNANAH FINAL MENU (2).xlsx')
ws2 = wb2.active

istik_items = {}
current_category = None
pending_description = None

for i, row in enumerate(ws2.iter_rows(values_only=True)):
    item_name = row[0]
    price = row[1] if len(row) > 1 else None
    
    if not item_name or 'ISTIKNANAH' in str(item_name):
        continue
    
    if item_name and not price:
        # Could be category
        if item_name in [
            'المشروبات الساخنه', 'عصائر فريش', 'مشروبات صحيه',
            'ايس درينك', 'ميلك شيك', 'مشروبات مميزة', 'صودا كوكتيل', 'سموزى'
        ]:
            current_category = item_name
            if current_category not in istik_items:
                istik_items[current_category] = []
            pending_description = None
        else:
            pending_description = item_name
    elif price and isinstance(price, (int, float)):
        if current_category:
            istik_items[current_category].append({
                'name': item_name,
                'price': int(price),
                'desc': pending_description or ''
            })
            pending_description = None

print('SOHO Items with Descriptions:')
for cat in list(soho_items.keys())[:3]:
    items = soho_items[cat]
    print(f"\n{cat} ({len(items)} items):")
    for item in items[:2]:
        print(f"  • {item['name']} ({item['price']} EGP)")
        if item['desc']:
            print(f"    > {item['desc'][:80]}...")

print('\n' + '='*100 + '\n')

print('ISTIKNANAH Items with Descriptions:')
for cat in list(istik_items.keys())[:3]:
    items = istik_items[cat]
    print(f"\n{cat} ({len(items)} items):")
    for item in items[:2]:
        print(f"  • {item['name']} ({item['price']} EGP)")
        if item['desc']:
            print(f"    > {item['desc'][:80]}...")

# Count totals
soho_count = sum(len(items) for items in soho_items.values())
istik_count = sum(len(items) for items in istik_items.values())
soho_with_desc = sum(1 for items in soho_items.values() for item in items if item.get('desc'))
istik_with_desc = sum(1 for items in istik_items.values() for item in items if item.get('desc'))

print(f'\n✓ Total SOHO items: {soho_count} ({soho_with_desc} with descriptions)')
print(f'✓ Total ISTIKNANAH items: {istik_count} ({istik_with_desc} with descriptions)')

# Save complete menu data
menu_data = {
    'soho': soho_items,
    'istiknanah': istik_items,
    'totals': {
        'soho': soho_count,
        'istiknanah': istik_count,
        'soho_with_desc': soho_with_desc,
        'istik_with_desc': istik_with_desc
    }
}

with open('menu_with_descriptions.json', 'w', encoding='utf-8') as f:
    json.dump(menu_data, f, ensure_ascii=False, indent=2)

print('\n✓ Saved complete menu to menu_with_descriptions.json')
