import openpyxl
import json

# Read Soho menu
print("=== SOHO MENU ===")
soho_wb = openpyxl.load_workbook('soho menu---.xlsx')
soho_sheet = soho_wb.active

for idx, row in enumerate(soho_sheet.iter_rows(min_row=1, max_row=100, values_only=True)):
    if any(row):
        print(f"Row {idx}: {row}")

print("\n=== ISTIKNANAH MENU ===")
istik_wb = openpyxl.load_workbook('ISTIKNANAH FINAL MENU (2).xlsx')
istik_sheet = istik_wb.active

for idx, row in enumerate(istik_sheet.iter_rows(min_row=1, max_row=100, values_only=True)):
    if any(row):
        print(f"Row {idx}: {row}")
