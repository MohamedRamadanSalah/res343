import openpyxl
import json

# Load SOHO menu from Excel
wb = openpyxl.load_workbook('soho menu---.xlsx')
ws = wb.active

soho_items = {}
current_category = None
for row in ws.iter_rows(values_only=True):
    item_name = row[0]
    price = row[1]
    
    if not item_name or not price:
        if item_name and not price:  # Category header
            current_category = item_name
            soho_items[current_category] = []
    else:
        if current_category and price:
            soho_items[current_category].append({
                'name': item_name,
                'price': price
            })

# Load ISTIKNANAH menu from Excel
wb2 = openpyxl.load_workbook('ISTIKNANAH FINAL MENU (2).xlsx')
ws2 = wb2.active

istik_items = {}
current_category = None
for row in ws2.iter_rows(values_only=True):
    item_name = row[0]
    price = row[1]
    
    if item_name and item_name.startswith('ISTIKNANAH'):
        continue
    
    if not item_name or not price:
        if item_name and not price:  # Category header
            current_category = item_name
            if current_category not in istik_items:
                istik_items[current_category] = []
    else:
        if current_category and price and isinstance(price, (int, float)):
            istik_items[current_category].append({
                'name': item_name,
                'price': price
            })

print('SOHO Categories and Items:')
for cat, items in soho_items.items():
    print(f"\n{cat}:")
    for item in items[:3]:  # Show first 3 items per category
        print(f"  - {item['name']} ({item['price']})")
    if len(items) > 3:
        print(f"  ... and {len(items) - 3} more")

print('\n' + '='*100 + '\n')

print('ISTIKNANAH Categories and Items:')
for cat, items in istik_items.items():
    if cat:  # Skip empty categories
        print(f"\n{cat}:")
        for item in items[:3]:  # Show first 3 items per category
            print(f"  - {item['name']} ({item['price']})")
        if len(items) > 3:
            print(f"  ... and {len(items) - 3} more")

# Save to JSON for reference
with open('soho_items.json', 'w', encoding='utf-8') as f:
    json.dump(soho_items, f, ensure_ascii=False, indent=2)

with open('istik_items.json', 'w', encoding='utf-8') as f:
    json.dump(istik_items, f, ensure_ascii=False, indent=2)

print('\n✓ Saved items to soho_items.json and istik_items.json')

# Count total items
soho_count = sum(len(items) for items in soho_items.values())
istik_count = sum(len(items) for items in istik_items.values())
print(f'\nTotal SOHO items: {soho_count}')
print(f'Total ISTIKNANAH items: {istik_count}')
