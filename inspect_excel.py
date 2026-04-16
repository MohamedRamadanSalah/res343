import openpyxl

# Read Soho menu 
soho_wb = openpyxl.load_workbook('soho menu---.xlsx')
soho_ws = soho_wb.active

print("First 30 rows of Soho menu:")
for i in range(1, 31):
    a = soho_ws.cell(i, 1).value
    b = soho_ws.cell(i, 2).value
    if a:
        a_str = str(a)[:60] if a else ''
        print(f"{i}: Name='{a_str}' Price='{b}'")
